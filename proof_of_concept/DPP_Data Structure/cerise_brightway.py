"""CE-RISE DPP -> Brightway bridge utilities.

This module provides:
- DPP JSON loading + schema/semantic validation
- Multi-DPP merge and secondary-flow resolution
- Brightway2-compatible foreground payload builder
- FU/reference-flow scaling from a study config JSON
- Excel template parser for visual user input
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Protocol, Tuple


def _runtime_default_bw_dir() -> str:
    explicit = os.environ.get("CERISE_BW_DIR") or os.environ.get("BW2_DIR") or os.environ.get("BRIGHTWAY2_DIR")
    if explicit:
        return explicit
    if Path("/var/data").exists():
        return "/var/data/brightway"
    return str(Path.home() / ".local" / "share" / "Brightway3")


_AUTO_BW_DIR = _runtime_default_bw_dir()
if _AUTO_BW_DIR:
    os.environ.setdefault("BRIGHTWAY2_DIR", _AUTO_BW_DIR)
    os.environ.setdefault("BW2_DIR", _AUTO_BW_DIR)


try:
    from jsonschema import Draft202012Validator
except Exception:  # pragma: no cover
    Draft202012Validator = None

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise RuntimeError("openpyxl is required to parse the Excel visual template") from exc

# Brightway imports are optional; module still works in stub mode without them
try:  # pragma: no cover
    import bw2calc as bc
except Exception:  # pragma: no cover
    bc = None
try:  # pragma: no cover
    import bw2data as bd
except Exception:  # pragma: no cover
    bd = None
try:  # pragma: no cover
    from bw2io.importers.bonsai import BonsaiImporter
    from bw2io.strategies.bonsai import mapb3
except Exception:  # pragma: no cover
    BonsaiImporter = None
    mapb3 = None
try:  # pragma: no cover
    from bw2io.remote import install_project
except Exception:  # pragma: no cover
    install_project = None


SCHEMA_PATH = Path(__file__).resolve().with_name("dpp.schema.json")
PROOF_OF_CONCEPT_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = PROOF_OF_CONCEPT_DIR.parent
DEFAULT_DB_NAME = "dpp_foreground"
DEFAULT_METHOD = ("EF v3.1", "climate change", "global warming potential (GWP100)")
DEFAULT_BW_DIR = Path(_AUTO_BW_DIR).expanduser()
DEFAULT_BW_PROJECT = "Bonsai_V3"
DEFAULT_REMOTE_PROJECT = "ecoinvent-3.10-biosphere"
DEFAULT_BONSAI_DB = "bonsai"
DEFAULT_BIOSPHERE_DB = "ecoinvent-3.10-biosphere"


def _env_path(*names: str) -> Optional[Path]:
    for name in names:
        raw = os.environ.get(name)
        if raw:
            return Path(raw).expanduser()
    return None


def _find_optional_repo_file(*names: str) -> Optional[Path]:
    for base in (PROOF_OF_CONCEPT_DIR, REPO_ROOT):
        for name in names:
            direct = base / name
            if direct.exists():
                return direct
            matches = sorted(base.rglob(name))
            if matches:
                return matches[0]
    return None


def _discover_bonsai_files_dir() -> Optional[Path]:
    env_dir = _env_path("CERISE_BONSAI_FILES_DIR", "BONSAI_FILES_DIR")
    if env_dir and (env_dir / "io_metadata.json").exists():
        return env_dir
    for candidate in (
        PROOF_OF_CONCEPT_DIR / "DPP_Data Structure" / "bonsai_files",
        PROOF_OF_CONCEPT_DIR / "bonsai_files",
        REPO_ROOT / "bonsai_files",
    ):
        if (candidate / "io_metadata.json").exists():
            return candidate
    for base in (PROOF_OF_CONCEPT_DIR, REPO_ROOT):
        matches = sorted(base.rglob("io_metadata.json"))
        if matches:
            return matches[0].parent
    return None


DEFAULT_CPA_MAPPING = _find_optional_repo_file("CPA_Bonsai_Mappin.xlsx")
DEFAULT_BONSAI_FILES = _discover_bonsai_files_dir()
DEFAULT_WORKBOOK_INPUT = (
    _env_path("CERISE_WORKBOOK_PATH", "CERISE_EXAMPLE_WORKBOOK")
    or _find_optional_repo_file(
        "Black ink printer cartridge manufacturing copy.xlsx",
        "Black ink printer cartridge manufacturing copy 2.xlsx",
    )
)
DEFAULT_EXAMPLE_DPP = Path(__file__).resolve().parent / "examples" / "dpp_from_black_ink_workbook.json"
DEFAULT_EXAMPLE_STUDY = Path(__file__).resolve().parent / "examples" / "study_config_from_wizard.json"
DEFAULT_EXAMPLE_FG = Path(__file__).resolve().parent / "examples" / "foreground_payload_from_dpp.json"

# Terminal UI colors (interactive mode only)
CLR_RESET = "\033[0m"
CLR_BOLD = "\033[1m"
CLR_TITLE = "\033[1;36m"
CLR_OPTION = "\033[1;34m"
CLR_PROMPT = "\033[1;33m"
CLR_INPUT = "\033[1;32m"
CLR_OK = "\033[1;32m"
CLR_WARN = "\033[1;33m"
CLR_ERR = "\033[1;31m"
CLR_DIM = "\033[2m"
BONSAI_CODE_RE = re.compile(r"^[A-Za-z0-9_]+\|[^|]+$")


@dataclass
class ResolvedKey:
    """Resolved provider key for technosphere linking."""

    database: str
    code: str

    def as_tuple(self) -> Tuple[str, str]:
        return (self.database, self.code)


class Resolver(Protocol):
    def resolve_secondary(
        self,
        flow_object_id: str,
        requesting_activity_id: str,
        flow: Optional[Dict[str, Any]] = None,
        flow_object: Optional[Dict[str, Any]] = None,
        activity: Optional[Dict[str, Any]] = None,
    ) -> Optional[ResolvedKey]:
        """Resolve a secondary flow object to a provider activity/database key."""


class BackgroundResolverStub:
    """Fallback resolver for background links.

    Parameters
    ----------
    mapping:
        Optional explicit map flowObjectId -> (database, code).
    default_db:
        Placeholder background DB name used when mapping is not provided.
    """

    def __init__(
        self,
        mapping: Optional[Dict[str, Tuple[str, str]]] = None,
        default_db: str = "background_stub",
    ) -> None:
        self.mapping = mapping or {}
        self.default_db = default_db

    def resolve_secondary(
        self,
        flow_object_id: str,
        requesting_activity_id: str,
        flow: Optional[Dict[str, Any]] = None,
        flow_object: Optional[Dict[str, Any]] = None,
        activity: Optional[Dict[str, Any]] = None,
    ) -> Optional[ResolvedKey]:
        if flow_object_id in self.mapping:
            db, code = self.mapping[flow_object_id]
            return ResolvedKey(db, code)
        return ResolvedKey(self.default_db, flow_object_id)


class MultiDPPResolver:
    """Resolve secondary inputs from combined DPP activities, else fallback to background."""

    def __init__(
        self,
        combined_dpp: Dict[str, Any],
        db_name: str = DEFAULT_DB_NAME,
        fallback: Optional[Resolver] = None,
    ) -> None:
        self.db_name = db_name
        self.fallback = fallback or BackgroundResolverStub()
        self.primary_map: Dict[str, ResolvedKey] = {}

        for activity in combined_dpp.get("activities", []):
            activity_id = activity["activityId"]
            det_id = activity["determiningFlowId"]
            det_flow = next((f for f in activity["flows"] if f["flowId"] == det_id), None)
            if not det_flow:
                continue
            self.primary_map[det_flow["flowObjectId"]] = ResolvedKey(self.db_name, activity_id)

    def resolve_secondary(
        self,
        flow_object_id: str,
        requesting_activity_id: str,
        flow: Optional[Dict[str, Any]] = None,
        flow_object: Optional[Dict[str, Any]] = None,
        activity: Optional[Dict[str, Any]] = None,
    ) -> Optional[ResolvedKey]:
        resolved = self.primary_map.get(flow_object_id)
        if resolved:
            return resolved
        return self.fallback.resolve_secondary(
            flow_object_id,
            requesting_activity_id,
            flow=flow,
            flow_object=flow_object,
            activity=activity,
        )


class BonsaiResolver:
    """Resolve secondary flow objects against BONSAI datasets in the active BW project."""

    def __init__(
        self,
        bonsai_db: str = DEFAULT_BONSAI_DB,
        cpa_mapping_path: Path | str | None = DEFAULT_CPA_MAPPING,
    ) -> None:
        if bd is None:
            raise RuntimeError("bw2data is not installed in this environment")
        if bonsai_db not in bd.databases:
            raise RuntimeError(f"BONSAI database '{bonsai_db}' not found in current project")
        self.bonsai_db = bonsai_db
        self.db = bd.Database(bonsai_db)
        # Kept in signature for CLI/API compatibility; strict resolver ignores inferred mappings.
        _ = cpa_mapping_path
        self._code_re = re.compile(r"^[AM]_(?P<stem>.+)\|(?P<loc>[A-Z]{2,3})$")
        self.name_to_stems: Dict[str, set[str]] = {}
        self.name_loc_codes: Dict[Tuple[str, str], List[str]] = {}
        for act in self.db:
            name_key = str(act.get("name") or "").strip().lower()
            code = str(act.get("code") or "")
            m = self._code_re.match(code)
            if not name_key or not m:
                continue
            stem = m.group("stem")
            loc = m.group("loc")
            self.name_to_stems.setdefault(name_key, set()).add(stem)
            self.name_loc_codes.setdefault((name_key, loc), []).append(code)

    def _resolve_by_code(self, code: str) -> Optional[ResolvedKey]:
        try:
            _ = self.db.get(code)
            return ResolvedKey(self.bonsai_db, code)
        except Exception:
            return None

    def resolve_secondary(
        self,
        flow_object_id: str,
        requesting_activity_id: str,
        flow: Optional[Dict[str, Any]] = None,
        flow_object: Optional[Dict[str, Any]] = None,
        activity: Optional[Dict[str, Any]] = None,
    ) -> Optional[ResolvedKey]:
        flow = flow or {}
        flow_object = flow_object or {}
        activity = activity or {}
        _ = (flow_object_id, requesting_activity_id)

        explicit_code = str(flow.get("bonsaiCode") or "").strip()
        if explicit_code:
            if not BONSAI_CODE_RE.match(explicit_code):
                raise ValueError(
                    f"Invalid bonsaiCode '{explicit_code}'; expected '<TOKEN>|<LOCATION>'"
                )
            resolved = self._resolve_by_code(explicit_code)
            if resolved is None:
                raise ValueError(f"BONSAI code '{explicit_code}' not found in '{self.bonsai_db}'")

            ds = self.db.get(explicit_code)
            src_loc = _normalize_location_code(flow.get("sourceLocation") or "")
            if src_loc:
                code_loc = _normalize_location_code(explicit_code.split("|", 1)[1])
                if code_loc != src_loc:
                    raise ValueError(
                        f"BONSAI code '{explicit_code}' location '{code_loc}' mismatches sourceLocation '{src_loc}'"
                    )

            bonsai_process = str(flow.get("bonsaiProcess") or "").strip()
            if bonsai_process:
                proc_norm = _norm_text(bonsai_process)
                name_norm = _norm_text(ds.get("name") or "")
                ref_norm = _norm_text(ds.get("reference product") or "")
                if proc_norm not in {name_norm, ref_norm}:
                    raise ValueError(
                        f"BONSAI code '{explicit_code}' does not match bonsaiProcess '{bonsai_process}'"
                    )
            return resolved

        # Explicit BONSAI code hints only (strict mode).
        for key in ("backgroundHint", "upstreamActivityRef", "classification", "cpa"):
            raw = flow.get(key) if key in flow else flow_object.get(key)
            if raw:
                txt = str(raw).strip().replace("BONSAI", "").strip()
                if re.match(r"^[AM]_[^|]+\\|[A-Z]{2,3}$", txt):
                    resolved = self._resolve_by_code(txt)
                    if resolved:
                        return resolved

        # Deterministic construction from classification category + location:
        # 1) Find BONSAI stems for the exact classification category name.
        # 2) Build M_<stem>|<LOC> first (market default), then A_<stem>|<LOC>.
        # 3) Use only if the constructed code exists in BONSAI.
        classification_name = str(
            flow.get("classification")
            or flow_object.get("classification")
            or flow.get("cpa")
            or flow_object.get("cpa")
            or ""
        ).strip().lower()
        if classification_name:
            flow_loc = str(flow.get("sourceLocation") or "").strip().upper()
            act_loc = str(((activity.get("location") or {}).get("countryCode") or "")).strip().upper()
            candidate_locs: List[str] = []
            for loc in (flow_loc, act_loc, "WE", "GLO"):
                if loc and loc not in candidate_locs:
                    candidate_locs.append(loc)

            stems = sorted(self.name_to_stems.get(classification_name, set()), key=str.lower)
            for loc in candidate_locs:
                for stem in stems:
                    for candidate in (f"M_{stem}|{loc}", f"A_{stem}|{loc}"):
                        resolved = self._resolve_by_code(candidate)
                        if resolved:
                            return resolved

        return None


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", s.strip().lower()).strip("-")


def _norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _normalize_location_code(code: str) -> str:
    raw = str(code or "").strip().upper()
    if not raw:
        return ""
    if raw == "GLOBAL":
        return "GLO"
    return raw


def set_brightway_context(
    project: str = DEFAULT_BW_PROJECT,
    bw_dir: Path | str = DEFAULT_BW_DIR,
) -> None:
    """Set Brightway data directory and current project."""
    if bd is None:
        raise RuntimeError("bw2data is not installed in this environment")
    os.environ["BRIGHTWAY2_DIR"] = str(bw_dir)
    os.environ["BW2_DIR"] = str(bw_dir)
    bd.projects.set_current(project)


def project_databases(project: str, bw_dir: Path | str = DEFAULT_BW_DIR) -> List[str]:
    """Return database names available in a Brightway project."""
    if bd is None:
        return []
    os.environ["BRIGHTWAY2_DIR"] = str(bw_dir)
    os.environ["BW2_DIR"] = str(bw_dir)
    bd.projects.set_current(project)
    return list(bd.databases)


def projects_with_database(db_name: str, bw_dir: Path | str = DEFAULT_BW_DIR) -> List[str]:
    """List Brightway projects which contain the given database name."""
    if bd is None:
        return []
    os.environ["BRIGHTWAY2_DIR"] = str(bw_dir)
    os.environ["BW2_DIR"] = str(bw_dir)
    found: List[str] = []
    current_obj = bd.projects.current
    current_name = getattr(current_obj, "name", None) if current_obj is not None else None
    for proj in bd.projects:
        name = getattr(proj, "name", None)
        if not name:
            s = str(proj)
            name = s.replace("Project: ", "", 1) if s.startswith("Project: ") else s
        bd.projects.set_current(name)
        if db_name in bd.databases:
            found.append(name)
    if current_name:
        bd.projects.set_current(current_name)
    return found


def import_bonsai_if_missing(
    project: str,
    bw_dir: Path | str = DEFAULT_BW_DIR,
    bonsai_db: str = DEFAULT_BONSAI_DB,
    bonsai_files: Path | str | None = DEFAULT_BONSAI_FILES,
    biosphere_db: str = DEFAULT_BIOSPHERE_DB,
    remote_project: str = DEFAULT_REMOTE_PROJECT,
) -> bool:
    """Import BONSAI technosphere into selected project if missing."""
    if bd is None:
        raise RuntimeError("bw2data is not installed in this environment")
    if BonsaiImporter is None or mapb3 is None:
        raise RuntimeError("bw2io BONSAI importer is not available in this environment")

    set_brightway_context(project=project, bw_dir=bw_dir)
    if biosphere_db not in bd.databases:
        if install_project is None:
            raise RuntimeError(
                f"target biosphere db missing ('{biosphere_db}') and bw2io.remote.install_project is unavailable"
            )
        # Seed empty/new projects with ecoinvent biosphere + methods.
        try:
            install_project(remote_project, project_name=project)
        except Exception as exc:
            msg = str(exc).lower()
            if "already exists" in msg:
                existing = set(bd.databases)
                safe_partial = {"bonsai biosphere"}
                if existing and not existing.issubset(safe_partial):
                    raise RuntimeError(
                        "Cannot auto-overwrite existing project to seed biosphere. "
                        f"Current databases: {sorted(existing)}. "
                        "Use an empty/new project, or manually install biosphere first."
                    )
                install_project(remote_project, project_name=project, overwrite_existing=True)
            else:
                raise
        bd.projects.set_current(project)
        if biosphere_db not in bd.databases:
            raise RuntimeError(
                f"target biosphere db missing ('{biosphere_db}') even after remote project install"
            )

    if bonsai_db in bd.databases:
        return False

    if bonsai_files is None:
        raise RuntimeError(
            "BONSAI files folder is not bundled in the current repo layout. "
            "Pass an explicit path or set CERISE_BONSAI_FILES_DIR / BONSAI_FILES_DIR."
        )

    files_path = Path(bonsai_files).expanduser()
    if not files_path.exists():
        raise RuntimeError(f"BONSAI files folder not found: {files_path}")
    if not (files_path / "io_metadata.json").exists():
        raise RuntimeError(
            f"BONSAI files folder is incomplete or incorrect: {files_path} "
            "(missing io_metadata.json)"
        )

    importer = BonsaiImporter(
        dirpath=files_path,
        db_name=bonsai_db,
        b3mapping=mapb3(),
    )
    importer.apply_strategies()
    importer.write_database()
    return True


def load_json(path: Path | str) -> Dict[str, Any]:
    p = Path(path)
    with p.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def save_json(path: Path | str, payload: Dict[str, Any]) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)


def load_schema(schema_path: Path | str = SCHEMA_PATH) -> Dict[str, Any]:
    return load_json(schema_path)


def validate_with_schema(instance: Dict[str, Any], schema: Dict[str, Any]) -> None:
    if Draft202012Validator is None:
        print("Warning: jsonschema package not installed; skipping JSON Schema validation.")
        return
    validator = Draft202012Validator(schema)
    errors = sorted(validator.iter_errors(instance), key=lambda e: list(e.absolute_path))
    if errors:
        lines = []
        for err in errors:
            path = "/".join(str(p) for p in err.absolute_path)
            lines.append(f"- {path or '<root>'}: {err.message}")
        raise ValueError("Schema validation failed:\n" + "\n".join(lines))


def validate_dpp_semantics(dpp: Dict[str, Any]) -> None:
    """Semantic checks that JSON Schema cannot fully express.

    Enforces:
    - determiningFlowId exists in each activity
    - determining flow is output
    - determining flow uses a primary flow object
    - no other flow in same activity uses a primary flow object
    """

    flow_objects = {fo["flowObjectId"]: fo for fo in dpp.get("flowObjects", [])}
    activity_ids = set()

    for activity in dpp.get("activities", []):
        activity_id = activity.get("activityId")
        if activity_id in activity_ids:
            raise ValueError(f"Duplicate activityId: {activity_id}")
        activity_ids.add(activity_id)

        flows = activity.get("flows", [])
        by_id = {f["flowId"]: f for f in flows}
        det_id = activity.get("determiningFlowId")
        if det_id not in by_id:
            raise ValueError(
                f"Activity '{activity_id}' determiningFlowId '{det_id}' not found in activity flows"
            )

        det_flow = by_id[det_id]
        if det_flow.get("direction") != "output":
            raise ValueError(
                f"Activity '{activity_id}' determining flow '{det_id}' must be an output"
            )

        det_obj_id = det_flow.get("flowObjectId")
        det_obj = flow_objects.get(det_obj_id)
        if det_obj is None:
            raise ValueError(
                f"Activity '{activity_id}' determining flow object '{det_obj_id}' not found"
            )
        if det_obj.get("objectClass") != "primary":
            raise ValueError(
                f"Activity '{activity_id}' determining flow object '{det_obj_id}' must be class 'primary'"
            )

        for flow in flows:
            fo = flow_objects.get(flow.get("flowObjectId"))
            if not fo:
                raise ValueError(
                    f"Activity '{activity_id}' references unknown flowObjectId '{flow.get('flowObjectId')}'"
                )
            if (
                fo.get("objectClass") == "primary"
                and flow["flowId"] != det_id
                and flow.get("direction") == "output"
            ):
                raise ValueError(
                    f"Activity '{activity_id}' has extra primary output flow '{flow['flowId']}'. "
                    "Primary output flows are only allowed as determining flow."
                )


def load_dpp_document(
    path: Path | str,
    schema_path: Path | str = SCHEMA_PATH,
    validate_schema: bool = True,
) -> Dict[str, Any]:
    dpp = load_json(path)
    if validate_schema:
        schema = load_schema(schema_path)
        validate_with_schema(dpp, schema)
    validate_dpp_semantics(dpp)
    return dpp


def combine_dpps(dpp_documents: Iterable[Dict[str, Any]]) -> Dict[str, Any]:
    docs = list(dpp_documents)
    if not docs:
        raise ValueError("No DPP documents provided")

    combined = {
        "schemaVersion": docs[0].get("schemaVersion", "1.0.0"),
        "dppId": "urn:dpp:combined:" + datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S"),
        "updatedAt": datetime.now(timezone.utc).isoformat(),
        "actors": [],
        "flowObjects": [],
        "activities": [],
    }

    actor_seen = set()
    flow_seen: Dict[str, Dict[str, Any]] = {}
    activity_seen = set()

    for doc in docs:
        for actor in doc.get("actors", []):
            key = (actor.get("actorId"), actor.get("role"))
            if key not in actor_seen:
                actor_seen.add(key)
                combined["actors"].append(actor)

        for fo in doc.get("flowObjects", []):
            fid = fo["flowObjectId"]
            existing = flow_seen.get(fid)
            if existing is None:
                flow_seen[fid] = dict(fo)
                continue

            # Merge duplicate flow objects from different DPPs.
            # Prefer "primary" if any contributor marks it as primary,
            # because it can be used as determining flow in another DPP.
            cls_existing = str(existing.get("objectClass", "")).lower()
            cls_new = str(fo.get("objectClass", "")).lower()
            if cls_existing != cls_new and "primary" in {cls_existing, cls_new}:
                existing["objectClass"] = "primary"

            # Fill missing optional metadata from the new occurrence.
            for key in ("name", "classification", "cpa", "ef31FlowId", "metadata"):
                if key not in existing and key in fo:
                    existing[key] = fo[key]

        for act in doc.get("activities", []):
            aid = act["activityId"]
            if aid in activity_seen:
                raise ValueError(f"Duplicate activityId across DPPs: {aid}")
            activity_seen.add(aid)
            combined["activities"].append(act)

    combined["flowObjects"] = list(flow_seen.values())
    validate_dpp_semantics(combined)
    return combined


class DPPBrightwayBuilder:
    """Convert combined DPP structure to a Brightway2 write payload.

    Output format is compatible with `Database.write(payload)` where keys are
    `(database, code)` tuples and values are dataset dictionaries.
    """

    def __init__(
        self,
        combined_dpp: Dict[str, Any],
        db_name: str = DEFAULT_DB_NAME,
        resolver: Optional[Resolver] = None,
        biosphere_db: str = DEFAULT_BIOSPHERE_DB,
        strict: bool = True,
    ) -> None:
        self.dpp = combined_dpp
        self.db_name = db_name
        self.flow_objects = {fo["flowObjectId"]: fo for fo in combined_dpp.get("flowObjects", [])}
        self.resolver = resolver or MultiDPPResolver(combined_dpp, db_name=db_name)
        self.strict = strict
        self.biosphere_db = biosphere_db
        self._biosphere_by_name: Dict[str, List[Any]] = {}
        if bd is not None and biosphere_db in getattr(bd, "databases", {}):
            for flow in bd.Database(biosphere_db):
                self._biosphere_by_name.setdefault(str(flow["name"]).strip().lower(), []).append(flow)

    def _resolve_biosphere_flow(self, flow_object: Dict[str, Any]) -> Optional[Tuple[str, str]]:
        if bd is None or not self._biosphere_by_name:
            return None
        name = str(flow_object.get("name") or "").strip().lower()
        if not name:
            return None
        candidates = self._biosphere_by_name.get(name, [])
        if not candidates:
            return None
        hit = candidates[0]
        return (self.biosphere_db, hit["code"])

    def build_foreground(self) -> Tuple[Dict[Tuple[str, str], Dict[str, Any]], List[Dict[str, str]]]:
        payload: Dict[Tuple[str, str], Dict[str, Any]] = {}
        unresolved: List[Dict[str, str]] = []

        for activity in self.dpp.get("activities", []):
            aid = activity["activityId"]
            det = next(f for f in activity["flows"] if f["flowId"] == activity["determiningFlowId"])
            det_fo = self.flow_objects[det["flowObjectId"]]

            exchanges: List[Dict[str, Any]] = [
                {
                    "input": (self.db_name, aid),
                    "type": "production",
                    "amount": float(det["amount"]),
                    "unit": det["unit"],
                }
            ]

            for flow in activity.get("flows", []):
                if flow["flowId"] == activity["determiningFlowId"]:
                    continue

                fo = self.flow_objects.get(flow["flowObjectId"])
                if not fo:
                    unresolved.append(
                        {
                            "activityId": aid,
                            "flowId": flow["flowId"],
                            "reason": "Unknown flowObjectId",
                        }
                    )
                    continue

                amount = float(flow["amount"])
                direction = flow["direction"]
                cls = fo["objectClass"]

                if cls in {"secondary", "primary"}:
                    provider = self.resolver.resolve_secondary(
                        fo["flowObjectId"],
                        aid,
                        flow=flow,
                        flow_object=fo,
                        activity=activity,
                    )
                    if provider is None:
                        flow_name = str(fo.get("name") or "<unnamed flow>")
                        classification = str(
                            flow.get("classification")
                            or fo.get("classification")
                            or flow.get("cpa")
                            or fo.get("cpa")
                            or "-"
                        )
                        upstream = str(flow.get("upstreamActivityRef") or "-")
                        src_loc = str(flow.get("sourceLocation") or "-")
                        unresolved.append(
                            {
                                "activityId": aid,
                                "flowId": flow["flowId"],
                                "reason": (
                                    f"Unresolved secondary flow '{flow_name}' "
                                    f"(classification={classification}, upstream={upstream}, sourceLocation={src_loc})"
                                ),
                            }
                        )
                        continue

                    ex_type = "technosphere"
                    ex_amount = amount
                    if direction == "output":
                        # For non-determining secondary outputs use substitution convention
                        ex_type = "substitution"
                        ex_amount = -abs(amount)

                    exchanges.append(
                        {
                            "input": provider.as_tuple(),
                            "type": ex_type,
                            "amount": ex_amount,
                            "unit": flow["unit"],
                        }
                    )

                elif cls == "elementary":
                    ex_amount = amount if direction == "output" else -amount
                    biosphere_input = self._resolve_biosphere_flow(fo)
                    if biosphere_input is None:
                        unresolved.append(
                            {
                                "activityId": aid,
                                "flowId": flow["flowId"],
                                "reason": f"Unresolved biosphere flow '{fo.get('name')}'",
                            }
                        )
                        continue
                    exchanges.append(
                        {
                            "input": biosphere_input,
                            "type": "biosphere",
                            "amount": ex_amount,
                            "unit": flow["unit"],
                        }
                    )

                else:
                    unresolved.append(
                        {
                            "activityId": aid,
                            "flowId": flow["flowId"],
                            "reason": f"Unsupported flow object class '{cls}'",
                        }
                    )

            dataset = {
                "name": activity["name"],
                "reference product": det_fo["name"],
                "unit": det["unit"],
                "location": (activity.get("location") or {}).get("countryCode", "GLO"),
                "type": "process",
                "comment": f"LC stage: {activity['lcStage']['label']}",
                "exchanges": exchanges,
            }
            payload[(self.db_name, aid)] = dataset

        if unresolved and self.strict:
            details = "\n".join(
                f"- {u['activityId']} / {u['flowId']}: {u['reason']}" for u in unresolved[:20]
            )
            raise RuntimeError(
                "Foreground build failed due to unresolved links/flows:\n"
                + details
                + ("\n..." if len(unresolved) > 20 else "")
            )
        return payload, unresolved


def compute_fu_scaling(
    combined_dpp: Dict[str, Any],
    study_config: Dict[str, Any],
    db_name: str = DEFAULT_DB_NAME,
) -> Tuple[float, Dict[Tuple[str, str], float], Dict[str, Any]]:
    """Compute demand scaling from reference-flow target.

    Scaling basis:
    - Preferred: `referenceFlowMapping.referenceFlow` (value + unit)
    - Backward-compatible fallback: `functionalUnit.howMuch` if referenceFlow is absent

    Unit behavior:
    - If target unit == determining flow unit, direct scaling is used.
    - If different, `referenceFlowMapping.unitConversionFactor` is required.
      This factor converts 1 determining-flow unit into target reference-flow unit.
      Example: determining unit = kg, target unit = g => factor = 1000.
    """

    activities = {a["activityId"]: a for a in combined_dpp.get("activities", [])}
    mapping = study_config["referenceFlowMapping"]
    fu = study_config.get("functionalUnit", {})

    activity_id = mapping["activityId"]
    flow_object_id = mapping["flowObjectId"]
    if activity_id not in activities:
        raise ValueError(f"Mapping activityId '{activity_id}' not found")

    activity = activities[activity_id]
    det = next((f for f in activity["flows"] if f["flowId"] == activity["determiningFlowId"]), None)
    if det is None:
        raise ValueError(f"Activity '{activity_id}' has no determining flow")

    if det["flowObjectId"] != flow_object_id:
        raise ValueError(
            f"Mapping flowObjectId '{flow_object_id}' does not match determining flow object "
            f"'{det['flowObjectId']}' for activity '{activity_id}'"
        )

    ref_target = mapping.get("referenceFlow")
    if isinstance(ref_target, dict) and "value" in ref_target and "unit" in ref_target:
        target_value = float(ref_target["value"])
        target_unit = str(ref_target["unit"])
        scaling_basis = "referenceFlow"
    else:
        # Backward compatibility with old configs
        target_value = float(fu["howMuch"]["value"])
        target_unit = fu["howMuch"]["unit"]
        scaling_basis = "functionalUnit.howMuch (legacy fallback)"
    det_amount = float(det["amount"])
    det_unit = det["unit"]

    if target_unit == det_unit:
        comparable_det = det_amount
    else:
        factor = mapping.get("unitConversionFactor")
        if factor is None:
            raise ValueError(
                f"Reference-flow unit '{target_unit}' differs from determining flow unit '{det_unit}'. "
                "Set referenceFlowMapping.unitConversionFactor in study config."
            )
        comparable_det = det_amount * float(factor)

    if comparable_det == 0:
        raise ValueError("Determining flow amount is zero; cannot scale to FU")

    scaling = target_value / comparable_det
    demand = {(db_name, activity_id): scaling}

    info = {
        "activityId": activity_id,
        "activityName": activity["name"],
        "determiningFlowId": det["flowId"],
        "determiningFlowObjectId": det["flowObjectId"],
        "determiningAmount": det_amount,
        "determiningUnit": det_unit,
        "fuValue": fu.get("howMuch", {}).get("value"),
        "fuUnit": fu.get("howMuch", {}).get("unit"),
        "referenceFlowValue": target_value,
        "referenceFlowUnit": target_unit,
        "scalingBasis": scaling_basis,
        "scalingFactor": scaling,
    }
    return scaling, demand, info


def write_foreground_database(
    payload: Dict[Tuple[str, str], Dict[str, Any]],
    db_name: str,
    overwrite: bool = True,
) -> None:
    """Write payload into active Brightway project database."""
    if bd is None:
        raise RuntimeError("bw2data is not installed in this environment")
    if db_name in bd.databases and overwrite:
        bd.Database(db_name).delete(warn=False)
    bd.Database(db_name).write(payload)


def run_lca(
    demand: Dict[Tuple[str, str], float],
    method: Any = DEFAULT_METHOD,
) -> Dict[str, Any]:
    """Run LCIA in Brightway."""
    if bc is None or bd is None:
        raise RuntimeError("Brightway runtime is not available in this environment")
    if method not in bd.methods:
        raise ValueError(f"LCIA method not found in current project: {method}")

    # Real Brightway execution path
    lca = bc.LCA(demand, method)
    lca.lci()
    lca.lcia()
    return {
        "status": "ok",
        "method": method,
        "score": lca.score,
    }


def _unit_maps_from_vocab(ws_vocab) -> Tuple[Dict[str, str], Dict[str, str]]:
    """Return (label->code, code->label) maps from Vocab sheet Units table."""
    label_to_code: Dict[str, str] = {}
    code_to_label: Dict[str, str] = {}

    in_units = False
    for row in range(1, ws_vocab.max_row + 1):
        a = ws_vocab.cell(row, 1).value
        b = ws_vocab.cell(row, 2).value
        if isinstance(a, str) and a.strip().lower() == "units":
            in_units = True
            continue
        if in_units:
            if isinstance(a, str) and a.strip().lower() == "flowtypes":
                break
            if not a or not b:
                continue
            code = str(a).strip().upper()
            label = str(b).strip()
            label_to_code[label.lower()] = code
            code_to_label[code] = label

    return label_to_code, code_to_label


def _stage_maps_from_vocab(ws_vocab) -> Dict[str, str]:
    """Return label->id map from Vocab Life Cycle Stage table."""
    out: Dict[str, str] = {}
    in_stage = False
    for row in range(1, ws_vocab.max_row + 1):
        a = ws_vocab.cell(row, 1).value
        b = ws_vocab.cell(row, 2).value
        if isinstance(a, str) and a.strip().lower() == "life cycle stage":
            in_stage = True
            continue
        if in_stage:
            if not a and not b:
                continue
            if isinstance(a, str) and a.strip().lower() == "code":
                continue
            if isinstance(a, str) and a.strip().lower() not in {
                "design_r&d",
                "manufacturing",
                "installation_distribution_retail",
                "use",
                "maintenance_repair_refurbishment",
                "end_of_life",
            }:
                continue
            if a and b:
                out[str(b).strip().lower()] = str(a).strip()
    return out


def _location_map(ws_location) -> Dict[str, str]:
    """Name->ISO code map from Location sheet."""
    out: Dict[str, str] = {}
    for row in range(2, ws_location.max_row + 1):
        code = ws_location.cell(row, 1).value
        name = ws_location.cell(row, 2).value
        if not code or not name:
            continue
        out[str(name).strip().lower()] = str(code).strip().upper()
    return out


def _to_iso_location(value: Any, name_to_iso: Dict[str, str]) -> str:
    if value is None:
        return "GLO"
    raw = str(value).strip()
    if not raw:
        return "GLO"
    up = raw.upper()
    if len(up) in (2, 3) and up.isalpha():
        return up
    return name_to_iso.get(raw.lower(), up)


def _unit_code(value: Any, label_to_code: Dict[str, str]) -> Optional[str]:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    up = raw.upper()
    if len(up) == 3 and up.isalnum():
        return up
    code = label_to_code.get(raw.lower())
    if code:
        return code

    # Graceful parsing for labels that include symbols/parentheses
    candidates = {
        "kg": "KGM",
        "gram": "GRM",
        "tonne": "TNE",
        "litre": "LTR",
        "cubic metre": "MTQ",
        "kilowatt hour": "KWH",
        "megajoule": "MJ",
        "kilometre": "KMT",
        "metre": "MTR",
        "square metre": "MTK",
        "tonne-kilometre": "TKM",
        "piece / unit": "C62",
    }
    low = raw.lower()
    for key, val in candidates.items():
        if key in low:
            return val
    return None


def _maybe_number(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except Exception:
        return None


def _is_blank(*values: Any) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return False
    return True


def _find_section_rows(ws_lci) -> Dict[str, int]:
    names = {
        "bom": ["bill of materials"],
        "manuf_inputs": ["manufacturing", "process inputs"],
        "manuf_outputs": ["manufacturing", "process outputs"],
        "manuf_packaging_inputs": ["manufacturing", "packaging", "inputs"],
        "use_inputs": ["use phase", "inputs"],
        "refurb_inputs": ["maintenance, repair, refurbishment", "process inputs"],
        "refurb_outputs": ["maintenance, repair, refurbishment", "process outputs"],
        "refurb_packaging_inputs": ["maintenance, repair, refurbishment", "packaging", "inputs"],
        "eol_inputs": ["end-of-life", "process inputs"],
        "eol_outputs": ["end-of-life", "process outputs"],
    }
    out: Dict[str, int] = {}
    for row in range(1, ws_lci.max_row + 1):
        val = ws_lci.cell(row, 1).value
        if not isinstance(val, str):
            continue
        v = val.strip().lower()
        for key, needles in names.items():
            if key in out:
                continue
            if all(n in v for n in needles):
                out[key] = row
    return out


def _find_value_by_label(ws_lci, label_contains: str, value_col: int = 2) -> Optional[Any]:
    needle = label_contains.strip().lower()
    for row in range(1, ws_lci.max_row + 1):
        left = ws_lci.cell(row, 1).value
        if isinstance(left, str) and needle in left.strip().lower():
            return ws_lci.cell(row, value_col).value
    return None


def _infer_reference_year(ws_lci) -> int:
    val = _find_value_by_label(ws_lci, "reference year")
    try:
        if val is not None:
            return int(float(val))
    except Exception:
        pass
    return datetime.now().year


def _infer_stage_label(ws_lci, valid_labels: Optional[set[str]] = None) -> str:
    # New template stores a stage hint in column J next to major section headers.
    for row in range(1, ws_lci.max_row + 1):
        stage = ws_lci.cell(row, 10).value
        if isinstance(stage, str) and stage.strip():
            candidate = stage.strip()
            if valid_labels:
                if candidate.lower() in valid_labels:
                    return candidate
            else:
                return candidate
    # Legacy template keeps it in B11.
    val = ws_lci["B11"].value
    return str(val).strip() if val else "Manufacturing"


def _section_bounds(section_rows: Dict[str, int], key: str, max_row: int) -> Tuple[int, int]:
    if key not in section_rows:
        return (0, -1)
    start = section_rows[key] + 2  # heading + header row
    current = section_rows[key]
    next_rows = sorted(r for r in section_rows.values() if r > current)
    end = (next_rows[0] - 1) if next_rows else max_row
    return start, end


def excel_to_dpp(
    workbook_path: Path | str,
    actor_id: str = "urn:actor:unknown",
    actor_role: str = "primaryDataReporter",
) -> Dict[str, Any]:
    """Parse the visual Excel interface and emit a chained DPP JSON object.

    Conversion rules implemented:
    - BoM section is always converted into an activity with LC stage
      "Raw material acquisition" and no location/reference year.
    - Other stages are converted into linked activities so the value chain is:
      Raw material -> Manufacturing -> (Distribution) -> Use ->
      (Distribution) -> Maintenance -> (Distribution) -> End-of-life.
    - Every activity has exactly one determining flow (PrimaryFlowObject output).
    - Secondary inputs are explicit flows (user-provided), and elementary flows
      are taken from output tables when marked as emissions/compartment outputs.
    """

    wb = load_workbook(workbook_path, data_only=False)
    ws_lci = wb["LCI_Input"]
    ws_vocab = wb["Vocab"]
    ws_location = wb["Location"]

    label_to_code, _ = _unit_maps_from_vocab(ws_vocab)
    stage_label_to_id = _stage_maps_from_vocab(ws_vocab)
    name_to_iso = _location_map(ws_location)
    section_rows = _find_section_rows(ws_lci)

    product_name = str(ws_lci["B4"].value or "Unnamed product").strip()
    global_reference_year = _infer_reference_year(ws_lci)

    ref_amount_raw = _find_value_by_label(ws_lci, "reference flow number")
    try:
        ref_amount = float(ref_amount_raw if ref_amount_raw is not None else 1.0)
    except Exception:
        ref_amount = 1.0

    declared_unit_raw = (
        _find_value_by_label(ws_lci, "declared unit value") or ws_lci["B7"].value or "C62"
    )
    declared_unit = _unit_code(declared_unit_raw, label_to_code) or str(declared_unit_raw)

    stage_id_defaults = {
        "raw material acquisition": "Raw_material_acquisition",
        "manufacturing": "Manufacturing",
        "installation/distribution/retail": "Installation_distribution_retail",
        "use": "Use",
        "maintenance, repair, refurbishment": "Maintenance_repair_refurbishment",
        "end-of-life": "End_of_life",
    }

    flow_objects: Dict[str, Dict[str, Any]] = {}
    activities: List[Dict[str, Any]] = []
    used_activity_ids: set[str] = set()

    def _parse_year(value: Any) -> Optional[int]:
        try:
            if value is None or (isinstance(value, str) and not value.strip()):
                return None
            return int(float(value))
        except Exception:
            return None

    def _value_near_label(
        anchor_row: int,
        label_contains: str,
        max_back: int = 20,
        max_fwd: int = 10,
        value_col: int = 2,
    ) -> Optional[Any]:
        needle = " ".join(label_contains.strip().lower().split())

        def _matches(cell_value: Any) -> bool:
            if not isinstance(cell_value, str):
                return False
            left = " ".join(cell_value.strip().lower().split())
            return left.startswith(needle)

        start = max(1, anchor_row - max_back)
        end = min(ws_lci.max_row, anchor_row + max_fwd)

        for row in range(anchor_row, start - 1, -1):
            left = ws_lci.cell(row, 1).value
            if _matches(left):
                return ws_lci.cell(row, value_col).value
        for row in range(anchor_row + 1, end + 1):
            left = ws_lci.cell(row, 1).value
            if _matches(left):
                return ws_lci.cell(row, value_col).value
        return None

    def _make_unique_activity_id(base_name: str) -> str:
        base = f"urn:activity:{_norm(base_name) or 'activity'}"
        if base not in used_activity_ids:
            used_activity_ids.add(base)
            return base
        i = 2
        while f"{base}-{i}" in used_activity_ids:
            i += 1
        out = f"{base}-{i}"
        used_activity_ids.add(out)
        return out

    def ensure_flow_object(
        name: str,
        object_class: str,
        classification: Optional[str] = None,
        ef31_flow_id: Optional[str] = None,
        key_suffix: Optional[str] = None,
    ) -> str:
        base = f"urn:flowobject:{object_class}:{_norm(name)}"
        key = f"{base}:{_norm(key_suffix)}" if key_suffix else base
        if key not in flow_objects:
            fo: Dict[str, Any] = {
                "flowObjectId": key,
                "name": name,
                "objectClass": object_class,
            }
            if classification:
                fo["classification"] = classification
            if ef31_flow_id:
                fo["ef31FlowId"] = ef31_flow_id
            flow_objects[key] = fo
        return key

    def _next_flow_id(activity: Dict[str, Any]) -> str:
        fid = f"f-{activity['__flowCounter']:04d}"
        activity["__flowCounter"] += 1
        return fid

    def _stage_payload(stage_label: str) -> Dict[str, str]:
        low = stage_label.strip().lower()
        stage_id = stage_label_to_id.get(low, stage_id_defaults.get(low, _norm(stage_label)))
        return {
            "id": stage_id or _norm(stage_label) or "stage",
            "label": stage_label,
        }

    def _build_activity(
        activity_type_label: str,
        stage_label: str,
        location_raw: Optional[Any],
        reference_year_raw: Optional[Any],
        primary_output_name: str,
    ) -> Dict[str, Any]:
        stage_label = str(stage_label).strip()
        activity_type = str(activity_type_label or stage_label).strip()

        is_raw_material_stage = stage_label.lower() == "raw material acquisition"
        loc_name: Optional[str] = None
        loc_code: Optional[str] = None
        if not _is_blank(location_raw):
            loc_name = str(location_raw).strip()
            loc_code = _to_iso_location(loc_name, name_to_iso)

        ref_year = _parse_year(reference_year_raw)
        if not is_raw_material_stage:
            if not loc_name:
                raise ValueError(
                    f"Missing Location for activity type '{activity_type}' (stage: {stage_label})"
                )
            if ref_year is None:
                raise ValueError(
                    f"Missing Reference year for activity type '{activity_type}' (stage: {stage_label})"
                )

        if is_raw_material_stage:
            full_name = activity_type
        else:
            full_name = activity_type
            if loc_name:
                full_name = f"{full_name}, {loc_name}"
            if ref_year is not None:
                full_name = f"{full_name} in {ref_year}"

        activity_id = _make_unique_activity_id(full_name)
        primary_flow_object_id = ensure_flow_object(
            primary_output_name,
            "primary",
            key_suffix=activity_id,
        )

        activity: Dict[str, Any] = {
            "activityId": activity_id,
            "identifier": full_name,  # dcterms:identifier-compatible text identifier
            "name": full_name,
            "activityType": {
                "id": _norm(activity_type) or "activity",
                "label": activity_type,
            },
            "lcStage": _stage_payload(stage_label),
            "location": None
            if is_raw_material_stage
            else {
                "countryCode": loc_code or "GLO",
                "countryName": loc_name or "",
            },
            "referenceYear": None if is_raw_material_stage else ref_year,
            "flows": [],
            "__flowCounter": 1,
            "__primaryFlowObjectId": primary_flow_object_id,
        }

        det_flow_id = _next_flow_id(activity)
        activity["determiningFlowId"] = det_flow_id
        activity["flows"].append(
            {
                "flowId": det_flow_id,
                "flowObjectId": primary_flow_object_id,
                "direction": "output",
                "amount": ref_amount,
                "unit": declared_unit,
                "isDetermining": True,
                "notes": "Determining flow (primary output) for this activity",
            }
        )
        return activity

    def _parse_input_section_into(
        activity: Dict[str, Any],
        key: str,
        has_flow_type_col: bool,
        default_flow_type: Optional[str] = None,
    ) -> None:
        start, end = _section_bounds(section_rows, key, ws_lci.max_row)
        if start == 0:
            return

        header_row = max(1, start - 1)
        headers = {
            c: str(ws_lci.cell(header_row, c).value or "").strip().lower() for c in range(1, 11)
        }

        for row in range(start, end + 1):
            if has_flow_type_col:
                flow_type = ws_lci.cell(row, 1).value
                description = ws_lci.cell(row, 2).value
                qty = ws_lci.cell(row, 3).value
                unit_val = ws_lci.cell(row, 4).value
                upstream = ws_lci.cell(row, 5).value
                classification = ws_lci.cell(row, 6).value
                extra_cols = range(7, 11)
            else:
                flow_type = default_flow_type
                description = ws_lci.cell(row, 1).value
                qty = ws_lci.cell(row, 2).value
                unit_val = ws_lci.cell(row, 3).value
                upstream = ws_lci.cell(row, 4).value
                classification = ws_lci.cell(row, 5).value
                extra_cols = range(6, 11)

            if _is_blank(description, qty, unit_val, flow_type, classification, upstream):
                continue

            amount = _maybe_number(qty)
            unit = _unit_code(unit_val, label_to_code)
            if description is None or amount is None or unit is None:
                continue

            classification_str = None if _is_blank(classification) else str(classification).strip()
            evidence: Optional[str] = None
            notes: Optional[str] = None
            supply_loc: Optional[str] = None

            for col in extra_cols:
                head = headers.get(col, "")
                val = ws_lci.cell(row, col).value
                if _is_blank(val):
                    continue
                if "evidence" in head:
                    evidence = str(val).strip().lower()
                elif "note" in head:
                    notes = str(val).strip()
                elif any(k in head for k in ("supply", "country", "location")):
                    supply_loc = _to_iso_location(val, name_to_iso)
                elif ("classification" in head or "cpa" in head) and not classification_str:
                    classification_str = str(val).strip()

            fo_id = ensure_flow_object(
                str(description).strip(),
                "secondary",
                classification=classification_str,
            )

            flow: Dict[str, Any] = {
                "flowId": _next_flow_id(activity),
                "flowObjectId": fo_id,
                "direction": "input",
                "amount": amount,
                "unit": unit,
            }
            if not _is_blank(upstream):
                up = str(upstream).strip()
                if up.lower() not in {"no", "none", "n/a"}:
                    flow["upstreamActivityRef"] = up
            if classification_str:
                flow["classification"] = classification_str
            if evidence:
                flow["evidenceMethod"] = evidence
            if notes:
                flow["notes"] = notes
            if supply_loc:
                flow["sourceLocation"] = supply_loc
            activity["flows"].append(flow)

    def _parse_output_section_into(activity: Dict[str, Any], key: str) -> None:
        start, end = _section_bounds(section_rows, key, ws_lci.max_row)
        if start == 0:
            return

        header_row = max(1, start - 1)
        headers = {
            c: str(ws_lci.cell(header_row, c).value or "").strip().lower() for c in range(1, 11)
        }

        for row in range(start, end + 1):
            flow_type = ws_lci.cell(row, 1).value
            description = ws_lci.cell(row, 2).value
            qty = ws_lci.cell(row, 3).value
            unit_val = ws_lci.cell(row, 4).value

            if _is_blank(flow_type, description, qty, unit_val):
                continue

            amount = _maybe_number(qty)
            unit = _unit_code(unit_val, label_to_code)
            if description is None or amount is None or unit is None:
                continue

            upstream: Optional[str] = None
            classification: Optional[str] = None
            evidence: Optional[str] = None
            notes: Optional[str] = None
            source_loc: Optional[str] = None
            compartment: Optional[str] = None
            waste_code: Optional[str] = None
            waste_op: Optional[str] = None

            for col in range(5, 11):
                head = headers.get(col, "")
                val = ws_lci.cell(row, col).value
                if _is_blank(val):
                    continue
                sval = str(val).strip()
                if "upstream" in head:
                    if sval.lower() not in {"no", "none", "n/a"}:
                        upstream = sval
                elif "classification" in head or "cpa" in head:
                    classification = sval
                elif any(k in head for k in ("supply", "country", "location")):
                    source_loc = _to_iso_location(sval, name_to_iso)
                elif "evidence" in head:
                    evidence = sval.lower()
                elif "note" in head:
                    notes = sval
                elif "compartment" in head:
                    compartment = sval
                elif "waste code" in head:
                    waste_code = sval
                elif "waste operation" in head:
                    waste_op = sval.upper()

            flow_type_norm = str(flow_type or "other").strip().lower()
            is_elementary = flow_type_norm == "emission" or not _is_blank(compartment)
            object_class = "elementary" if is_elementary else "secondary"
            ef31_id = (
                f"urn:ef31:placeholder:{_norm(str(description))}" if is_elementary else None
            )

            fo_id = ensure_flow_object(
                str(description).strip(),
                object_class,
                classification=classification,
                ef31_flow_id=ef31_id,
            )

            flow: Dict[str, Any] = {
                "flowId": _next_flow_id(activity),
                "flowObjectId": fo_id,
                "direction": "output",
                "amount": amount,
                "unit": unit,
            }
            if upstream:
                flow["upstreamActivityRef"] = upstream
            if classification:
                flow["classification"] = classification
            if source_loc:
                flow["sourceLocation"] = source_loc
            if evidence:
                flow["evidenceMethod"] = evidence
            if notes:
                flow["notes"] = notes
            if compartment:
                flow["emissionCompartment"] = compartment
            if waste_code:
                flow["wasteCode"] = waste_code
            if waste_op:
                flow["wasteOperation"] = waste_op
            activity["flows"].append(flow)

    def _add_chain_input(activity: Dict[str, Any], previous_activity: Dict[str, Any]) -> None:
        activity["flows"].append(
            {
                "flowId": _next_flow_id(activity),
                "flowObjectId": previous_activity["__primaryFlowObjectId"],
                "direction": "input",
                "amount": ref_amount,
                "unit": declared_unit,
                "upstreamActivityRef": previous_activity["activityId"],
                "notes": "Linked to previous activity in product value chain",
            }
        )

    def _section_has_payload(key: str, has_flow_type_col: bool) -> bool:
        start, end = _section_bounds(section_rows, key, ws_lci.max_row)
        if start == 0:
            return False
        for row in range(start, end + 1):
            if has_flow_type_col:
                desc = ws_lci.cell(row, 2).value
                qty = ws_lci.cell(row, 3).value
                unit_val = ws_lci.cell(row, 4).value
            else:
                desc = ws_lci.cell(row, 1).value
                qty = ws_lci.cell(row, 2).value
                unit_val = ws_lci.cell(row, 3).value
            if _is_blank(desc, qty, unit_val):
                continue
            if _maybe_number(qty) is None:
                continue
            if _unit_code(unit_val, label_to_code) is None:
                continue
            return True
        return False

    def _extract_stage_context(
        stage_name: str,
        anchor_row: int,
        location_label: str,
        back: int,
        fwd: int,
        default_activity_description: Optional[str] = None,
    ) -> Optional[Tuple[str, Any, Any]]:
        desc = _value_near_label(anchor_row, "activity description", max_back=back, max_fwd=fwd)
        year = _value_near_label(anchor_row, "reference year", max_back=back, max_fwd=fwd)
        loc = _value_near_label(anchor_row, location_label, max_back=back, max_fwd=fwd)
        has_any = not _is_blank(desc, year, loc)
        if not has_any:
            return None
        if _is_blank(desc):
            if default_activity_description:
                desc = default_activity_description
            else:
                raise ValueError(
                    f"{stage_name} context is incomplete: missing Activity description"
                )
        missing = []
        if _is_blank(year):
            missing.append("Reference year")
        if _is_blank(loc):
            missing.append("Location")
        if missing:
            raise ValueError(
                f"{stage_name} context is incomplete: missing {', '.join(missing)}"
            )
        return str(desc).strip(), year, loc

    # ----- Stage/activity metadata extraction (strict, no auto-solvers) -----
    manuf_anchor = section_rows.get("manuf_inputs", 1)
    use_anchor = section_rows.get("use_inputs", manuf_anchor)
    refurb_anchor = section_rows.get("refurb_inputs", use_anchor)
    eol_anchor = section_rows.get("eol_inputs", refurb_anchor)

    bom_has_data = _section_has_payload("bom", has_flow_type_col=False)
    manuf_has_data = (
        _section_has_payload("manuf_inputs", has_flow_type_col=True)
        or _section_has_payload("manuf_outputs", has_flow_type_col=True)
        or _section_has_payload("manuf_packaging_inputs", has_flow_type_col=False)
    )
    use_has_data = _section_has_payload("use_inputs", has_flow_type_col=True)
    refurb_has_data = (
        _section_has_payload("refurb_inputs", has_flow_type_col=True)
        or _section_has_payload("refurb_outputs", has_flow_type_col=True)
        or _section_has_payload("refurb_packaging_inputs", has_flow_type_col=False)
    )
    eol_has_data = (
        _section_has_payload("eol_inputs", has_flow_type_col=True)
        or _section_has_payload("eol_outputs", has_flow_type_col=True)
    )

    if not bom_has_data:
        raise ValueError(
            "BoM section has no valid inputs (description, numeric quantity, valid unit)."
        )
    if not manuf_has_data:
        raise ValueError(
            "Manufacturing section has no valid inputs/outputs. "
            "At least one manufacturing flow is required."
        )

    manuf_ctx = _extract_stage_context(
        stage_name="Manufacturing",
        anchor_row=manuf_anchor,
        location_label="activity location",
        back=4,
        fwd=12,
    )
    if manuf_ctx is None:
        raise ValueError(
            "Manufacturing context is missing. Provide Activity description, "
            "Reference year, and Activity Location."
        )
    manuf_type, manuf_year, manuf_loc = manuf_ctx

    use_ctx = _extract_stage_context(
        stage_name="Use",
        anchor_row=use_anchor,
        location_label="use location",
        back=8,
        fwd=12,
        default_activity_description="Use phase",
    )
    refurb_ctx = _extract_stage_context(
        stage_name="Maintenance, repair, refurbishment",
        anchor_row=refurb_anchor,
        location_label="activity location",
        back=25,
        fwd=10,
        default_activity_description="Maintenance, repair, refurbishment",
    ) or _extract_stage_context(
        stage_name="Maintenance, repair, refurbishment",
        anchor_row=refurb_anchor,
        location_label="refurb location",
        back=25,
        fwd=10,
        default_activity_description="Maintenance, repair, refurbishment",
    )
    eol_ctx = _extract_stage_context(
        stage_name="End-of-life",
        anchor_row=eol_anchor,
        location_label="activity location",
        back=25,
        fwd=10,
        default_activity_description="End-of-life",
    ) or _extract_stage_context(
        stage_name="End-of-life",
        anchor_row=eol_anchor,
        location_label="eol location",
        back=25,
        fwd=10,
        default_activity_description="End-of-life",
    )

    if use_has_data and use_ctx is None:
        raise ValueError(
            "Use stage has flows but missing context. Provide Use location and Reference year "
            "(and optional Activity description)."
        )
    if refurb_has_data and refurb_ctx is None:
        raise ValueError(
            "Maintenance/refurbishment stage has flows but missing context. "
            "Provide Activity description, Reference year, and Activity Location (or Refurb location)."
        )
    if eol_has_data and eol_ctx is None:
        raise ValueError(
            "End-of-life stage has flows but missing context. "
            "Provide Activity description, Reference year, and Activity Location (or EoL location)."
        )

    # ----- Build chained activities (compact strict version) -----
    raw_material_activity = _build_activity(
        activity_type_label=f"{product_name} - BoM",
        stage_label="Raw material acquisition",
        location_raw=None,
        reference_year_raw=None,
        primary_output_name=f"{product_name} raw material package",
    )
    _parse_input_section_into(
        raw_material_activity,
        "bom",
        has_flow_type_col=False,
        default_flow_type="material",
    )
    activities.append(raw_material_activity)
    previous = raw_material_activity

    manufacturing_activity = _build_activity(
        activity_type_label=manuf_type,
        stage_label="Manufacturing",
        location_raw=manuf_loc,
        reference_year_raw=manuf_year,
        primary_output_name=f"{product_name} manufactured cartridge",
    )
    _add_chain_input(manufacturing_activity, previous)
    _parse_input_section_into(manufacturing_activity, "manuf_inputs", has_flow_type_col=True)
    _parse_output_section_into(manufacturing_activity, "manuf_outputs")
    _parse_input_section_into(
        manufacturing_activity,
        "manuf_packaging_inputs",
        has_flow_type_col=False,
        default_flow_type="packaging",
    )
    activities.append(manufacturing_activity)
    previous = manufacturing_activity

    if use_ctx is not None:
        use_type, use_year, use_loc = use_ctx
        use_activity = _build_activity(
            activity_type_label=use_type,
            stage_label="Use",
            location_raw=use_loc,
            reference_year_raw=use_year,
            primary_output_name=f"{product_name} used cartridge",
        )
        _add_chain_input(use_activity, previous)
        _parse_input_section_into(use_activity, "use_inputs", has_flow_type_col=True)
        activities.append(use_activity)
        previous = use_activity

    if refurb_ctx is not None:
        refurb_type, refurb_year, refurb_loc = refurb_ctx
        refurb_activity = _build_activity(
            activity_type_label=refurb_type,
            stage_label="Maintenance, repair, refurbishment",
            location_raw=refurb_loc,
            reference_year_raw=refurb_year,
            primary_output_name=f"{product_name} refilled cartridge",
        )
        _add_chain_input(refurb_activity, previous)
        _parse_input_section_into(refurb_activity, "refurb_inputs", has_flow_type_col=True)
        _parse_output_section_into(refurb_activity, "refurb_outputs")
        _parse_input_section_into(
            refurb_activity,
            "refurb_packaging_inputs",
            has_flow_type_col=False,
            default_flow_type="packaging",
        )
        activities.append(refurb_activity)
        previous = refurb_activity

    if eol_ctx is not None:
        eol_type, eol_year, eol_loc = eol_ctx
        eol_activity = _build_activity(
            activity_type_label=eol_type,
            stage_label="End-of-life",
            location_raw=eol_loc,
            reference_year_raw=eol_year,
            primary_output_name=f"{product_name} end-of-life treatment",
        )
        _add_chain_input(eol_activity, previous)
        _parse_input_section_into(eol_activity, "eol_inputs", has_flow_type_col=True)
        _parse_output_section_into(eol_activity, "eol_outputs")
        activities.append(eol_activity)

    # Remove internal builder keys
    for act in activities:
        act.pop("__flowCounter", None)
        act.pop("__primaryFlowObjectId", None)

    dpp_loc = _to_iso_location(manuf_loc, name_to_iso) if not _is_blank(manuf_loc) else "MULTI"
    dpp_id = f"urn:dpp:{_norm(product_name)}:{global_reference_year}:{_norm(dpp_loc)}"

    dpp = {
        "schemaVersion": "1.0.0",
        "dppId": dpp_id,
        "updatedAt": datetime.now(timezone.utc).isoformat(),
        "actors": [
            {
                "actorId": actor_id,
                "role": actor_role,
            }
        ],
        "flowObjects": list(flow_objects.values()),
        "activities": activities,
    }
    return dpp


def interactive_study_config_from_dpp(dpp: Dict[str, Any]) -> Dict[str, Any]:
    """Prompt user for FU + reference flow mapping from current DPP."""

    activities = dpp.get("activities", [])
    if not activities:
        raise ValueError("DPP has no activities")

    dpp_id = dpp.get("dppId", "<unknown dpp>")
    print(f"\nDPP selected: {dpp_id}")
    if len(activities) == 1:
        act = activities[0]
        print(f"Using activity: {act['name']} ({act['activityId']})")
    else:
        print("Available activities:")
        for a in activities:
            print(f"  - {a['name']} ({a['activityId']})")
        default_key = activities[0]["activityId"]
        selected = input(
            f"Functional provider activity (name or activityId) [default: {default_key}]: "
        ).strip()
        if not selected:
            act = activities[0]
        else:
            selected_l = selected.lower()
            act = next(
                (
                    a
                    for a in activities
                    if selected_l == str(a.get("activityId", "")).lower()
                    or selected_l == str(a.get("name", "")).lower()
                ),
                None,
            )
            if act is None:
                raise ValueError(
                    f"Activity '{selected}' not found in DPP. "
                    "Use exact activity name or activityId."
                )

    det = next(f for f in act["flows"] if f["flowId"] == act["determiningFlowId"])
    flow_object_id = det["flowObjectId"]

    print("\nFunctional unit definition (metadata for comparability, not used for scaling)")
    what_default = next(
        (fo["name"] for fo in dpp["flowObjects"] if fo["flowObjectId"] == flow_object_id),
        act["name"],
    )
    what = input(f"What [default: {what_default}]: ").strip() or what_default

    val_default = str(det["amount"])
    val_text = input(f"How much (numeric) [default: {val_default}]: ").strip() or val_default
    how_much_value = float(val_text)

    unit_default = det["unit"]
    how_much_unit = input(f"How much unit [default: {unit_default}]: ").strip() or unit_default

    how_well = input("How well [default: as specified]: ").strip() or "as specified"
    how_long = input("How long [default: not time-dependent]: ").strip() or "not time-dependent"

    print("\nReference flow target (used for scaling)")
    rf_value_text = input(
        f"Reference flow amount (numeric) [default: {det['amount']}]: "
    ).strip() or str(det["amount"])
    rf_value = float(rf_value_text)
    rf_unit = input(
        f"Reference flow unit [default: {det['unit']}]: "
    ).strip() or det["unit"]

    mapping: Dict[str, Any] = {
        "activityId": act["activityId"],
        "flowObjectId": flow_object_id,
        "referenceFlow": {
            "value": rf_value,
            "unit": rf_unit,
        },
    }

    if rf_unit != det["unit"]:
        print(
            f"\nUnits differ: determining flow is {det['unit']} but reference flow is {rf_unit}."
        )
        factor_txt = input(
            "Enter unitConversionFactor (1 determining unit -> reference-flow unit): "
        ).strip()
        if not factor_txt:
            raise ValueError("unitConversionFactor is required when units differ")
        mapping["unitConversionFactor"] = float(factor_txt)

    study = {
        "functionalUnit": {
            "what": what,
            "howMuch": {
                "value": how_much_value,
                "unit": how_much_unit,
            },
            "howWell": how_well,
            "howLong": how_long,
        },
        "referenceFlowMapping": mapping,
    }
    return study


def run_end_to_end(
    workbook_path: Path,
    dpp_out: Path,
    study_out: Path,
    schema_path: Path = SCHEMA_PATH,
    project: str = DEFAULT_BW_PROJECT,
    bw_dir: Path | str = DEFAULT_BW_DIR,
    db_name: str = DEFAULT_DB_NAME,
    bonsai_db: str = DEFAULT_BONSAI_DB,
    biosphere_db: str = DEFAULT_BIOSPHERE_DB,
    method: Any = DEFAULT_METHOD,
) -> Dict[str, Any]:
    """Excel visual input -> strict pre-check -> DPP JSON -> FU wizard -> scaling + BW payload."""

    set_brightway_context(project=project, bw_dir=bw_dir)
    dpp = excel_to_dpp(workbook_path)
    schema = load_schema(schema_path)
    validate_with_schema(dpp, schema)
    validate_dpp_semantics(dpp)
    payload, unresolved = _build_foreground_strict_payload(
        dpp=dpp,
        db_name=db_name,
        bonsai_db=bonsai_db,
        biosphere_db=biosphere_db,
    )
    save_json(dpp_out, dpp)

    print(f"\nDPP JSON created: {dpp_out}")
    study = interactive_study_config_from_dpp(dpp)
    save_json(study_out, study)
    print(f"Study config saved: {study_out}")

    scaling, demand, info = compute_fu_scaling(dpp, study, db_name=db_name)
    write_foreground_database(payload, db_name=db_name, overwrite=True)

    print("\nScaling result")
    print(f"  Activity: {info['activityName']}")
    print(f"  Determining flow amount: {info['determiningAmount']} {info['determiningUnit']}")
    print(f"  FU: {info['fuValue']} {info['fuUnit']}")
    print(f"  Required activity units for FU: {scaling:.6g}")

    lcia_result = run_lca(demand, method=method)

    return {
        "dppPath": str(dpp_out),
        "studyPath": str(study_out),
        "scaling": scaling,
        "demand": {str(k): v for k, v in demand.items()},
        "unresolved": unresolved,
        "foregroundDatasetCount": len(payload),
        "lcia": lcia_result,
    }


def _build_foreground_strict_payload(
    dpp: Dict[str, Any],
    db_name: str,
    bonsai_db: str,
    biosphere_db: str,
) -> Tuple[Dict[Tuple[str, str], Dict[str, Any]], List[Dict[str, Any]]]:
    """Build strict foreground payload and raise on unresolved links."""
    fallback = BonsaiResolver(bonsai_db=bonsai_db, cpa_mapping_path=DEFAULT_CPA_MAPPING)
    resolver = MultiDPPResolver(dpp, db_name=db_name, fallback=fallback)
    builder = DPPBrightwayBuilder(
        dpp,
        db_name=db_name,
        resolver=resolver,
        biosphere_db=biosphere_db,
        strict=True,
    )
    return builder.build_foreground()


def _cmd_excel_to_dpp(args: argparse.Namespace) -> None:
    dpp = excel_to_dpp(args.workbook, actor_id=args.actor_id, actor_role=args.actor_role)
    schema = load_schema(args.schema)
    validate_with_schema(dpp, schema)
    validate_dpp_semantics(dpp)
    save_json(args.output, dpp)
    print(f"DPP written to: {args.output}")


def _cmd_combine(args: argparse.Namespace) -> None:
    docs = [load_dpp_document(p, schema_path=args.schema, validate_schema=True) for p in args.dpp]
    combined = combine_dpps(docs)
    save_json(args.output, combined)
    print(f"Combined DPP written to: {args.output}")


def _cmd_scale(args: argparse.Namespace) -> None:
    set_brightway_context(project=args.project, bw_dir=args.bw_dir)
    dpp = load_dpp_document(args.dpp, schema_path=args.schema, validate_schema=True)
    study = load_json(args.study)
    scaling, demand, info = compute_fu_scaling(dpp, study, db_name=args.db_name)
    print(json.dumps({"scaling": scaling, "demand": {str(k): v for k, v in demand.items()}, "info": info}, indent=2))


def _cmd_foreground(args: argparse.Namespace) -> None:
    set_brightway_context(project=args.project, bw_dir=args.bw_dir)
    dpp = load_dpp_document(args.dpp, schema_path=args.schema, validate_schema=True)
    fallback = BonsaiResolver(bonsai_db=args.bonsai_db, cpa_mapping_path=args.cpa_mapping)
    resolver = MultiDPPResolver(dpp, db_name=args.db_name, fallback=fallback)
    builder = DPPBrightwayBuilder(
        dpp,
        db_name=args.db_name,
        resolver=resolver,
        biosphere_db=args.biosphere_db,
        strict=not args.allow_unresolved,
    )
    payload, unresolved = builder.build_foreground()
    if args.write_bw:
        write_foreground_database(payload, db_name=args.db_name, overwrite=True)
        print(f"Foreground DB written to Brightway project '{args.project}': {args.db_name}")

    serializable = {f"{k[0]}::{k[1]}": v for k, v in payload.items()}
    out = {
        "database": args.db_name,
        "datasets": serializable,
        "unresolved": unresolved,
    }
    save_json(args.output, out)
    print(f"Foreground payload written to: {args.output}")
    if unresolved:
        print(f"Unresolved links: {len(unresolved)}")


def _cmd_run_lca(args: argparse.Namespace) -> None:
    set_brightway_context(project=args.project, bw_dir=args.bw_dir)
    dpp = load_dpp_document(args.dpp, schema_path=args.schema, validate_schema=True)
    study = load_json(args.study)
    scaling, demand, info = compute_fu_scaling(dpp, study, db_name=args.db_name)
    fallback = BonsaiResolver(bonsai_db=args.bonsai_db, cpa_mapping_path=args.cpa_mapping)
    resolver = MultiDPPResolver(dpp, db_name=args.db_name, fallback=fallback)
    builder = DPPBrightwayBuilder(
        dpp,
        db_name=args.db_name,
        resolver=resolver,
        biosphere_db=args.biosphere_db,
        strict=True,
    )
    payload, unresolved = builder.build_foreground()
    write_foreground_database(payload, db_name=args.db_name, overwrite=True)
    result = run_lca(demand, method=tuple(args.method))
    out = {
        "project": args.project,
        "database": args.db_name,
        "scaling": scaling,
        "demand": {str(k): v for k, v in demand.items()},
        "mappingInfo": info,
        "unresolved": unresolved,
        "lcia": result,
    }
    print(json.dumps(out, indent=2))


def _cmd_end_to_end(args: argparse.Namespace) -> None:
    result = run_end_to_end(
        workbook_path=args.workbook,
        dpp_out=args.dpp_out,
        study_out=args.study_out,
        schema_path=args.schema,
        project=args.project,
        bw_dir=args.bw_dir,
        db_name=args.db_name,
        bonsai_db=args.bonsai_db,
        biosphere_db=args.biosphere_db,
        method=tuple(args.method),
    )
    print("\nRun summary")
    print(json.dumps(result, indent=2))


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="CE-RISE DPP/Brightway bridge")
    sub = parser.add_subparsers(dest="cmd", required=False)

    p_excel = sub.add_parser("excel-to-dpp", help="Convert visual Excel input into DPP JSON")
    p_excel.add_argument("--workbook", type=Path, required=True)
    p_excel.add_argument("--output", type=Path, required=True)
    p_excel.add_argument("--schema", type=Path, default=SCHEMA_PATH)
    p_excel.add_argument("--actor-id", default="urn:actor:unknown")
    p_excel.add_argument("--actor-role", default="primaryDataReporter")
    p_excel.set_defaults(func=_cmd_excel_to_dpp)

    p_combine = sub.add_parser("combine", help="Combine multiple DPP documents")
    p_combine.add_argument("--dpp", type=Path, nargs="+", required=True)
    p_combine.add_argument("--output", type=Path, required=True)
    p_combine.add_argument("--schema", type=Path, default=SCHEMA_PATH)
    p_combine.set_defaults(func=_cmd_combine)

    p_scale = sub.add_parser("scale", help="Compute FU scaling and demand from study config")
    p_scale.add_argument("--dpp", type=Path, required=True)
    p_scale.add_argument("--study", type=Path, required=True)
    p_scale.add_argument("--schema", type=Path, default=SCHEMA_PATH)
    p_scale.add_argument("--db-name", default=DEFAULT_DB_NAME)
    p_scale.add_argument("--project", default=DEFAULT_BW_PROJECT)
    p_scale.add_argument("--bw-dir", type=Path, default=DEFAULT_BW_DIR)
    p_scale.set_defaults(func=_cmd_scale)

    p_fg = sub.add_parser("build-foreground", help="Build Brightway-compatible foreground payload")
    p_fg.add_argument("--dpp", type=Path, required=True)
    p_fg.add_argument("--output", type=Path, required=True)
    p_fg.add_argument("--schema", type=Path, default=SCHEMA_PATH)
    p_fg.add_argument("--db-name", default=DEFAULT_DB_NAME)
    p_fg.add_argument("--project", default=DEFAULT_BW_PROJECT)
    p_fg.add_argument("--bw-dir", type=Path, default=DEFAULT_BW_DIR)
    p_fg.add_argument("--bonsai-db", default=DEFAULT_BONSAI_DB)
    p_fg.add_argument("--biosphere-db", default=DEFAULT_BIOSPHERE_DB)
    p_fg.add_argument("--cpa-mapping", type=Path, default=DEFAULT_CPA_MAPPING)
    p_fg.add_argument("--write-bw", action="store_true", help="Also write payload into Brightway DB")
    p_fg.add_argument("--allow-unresolved", action="store_true", help="Do not fail on unresolved links")
    p_fg.set_defaults(func=_cmd_foreground)

    p_lca = sub.add_parser("run-lca", help="Build/write foreground and run real LCIA in Brightway")
    p_lca.add_argument("--dpp", type=Path, required=True)
    p_lca.add_argument("--study", type=Path, required=True)
    p_lca.add_argument("--schema", type=Path, default=SCHEMA_PATH)
    p_lca.add_argument("--db-name", default=DEFAULT_DB_NAME)
    p_lca.add_argument("--project", default=DEFAULT_BW_PROJECT)
    p_lca.add_argument("--bw-dir", type=Path, default=DEFAULT_BW_DIR)
    p_lca.add_argument("--bonsai-db", default=DEFAULT_BONSAI_DB)
    p_lca.add_argument("--biosphere-db", default=DEFAULT_BIOSPHERE_DB)
    p_lca.add_argument("--cpa-mapping", type=Path, default=DEFAULT_CPA_MAPPING)
    p_lca.add_argument(
        "--method",
        nargs=3,
        metavar=("M0", "M1", "M2"),
        default=list(DEFAULT_METHOD),
        help="LCIA method tuple, e.g. 'EF v3.1' 'climate change' 'global warming potential (GWP100)'",
    )
    p_lca.set_defaults(func=_cmd_run_lca)

    p_e2e = sub.add_parser("end-to-end", help="Excel -> DPP -> FU wizard -> scaling")
    p_e2e.add_argument("--workbook", type=Path, required=True)
    p_e2e.add_argument("--dpp-out", type=Path, required=True)
    p_e2e.add_argument("--study-out", type=Path, required=True)
    p_e2e.add_argument("--schema", type=Path, default=SCHEMA_PATH)
    p_e2e.add_argument("--project", default=DEFAULT_BW_PROJECT)
    p_e2e.add_argument("--bw-dir", type=Path, default=DEFAULT_BW_DIR)
    p_e2e.add_argument("--db-name", default=DEFAULT_DB_NAME)
    p_e2e.add_argument("--bonsai-db", default=DEFAULT_BONSAI_DB)
    p_e2e.add_argument("--biosphere-db", default=DEFAULT_BIOSPHERE_DB)
    p_e2e.add_argument(
        "--method",
        nargs=3,
        metavar=("M0", "M1", "M2"),
        default=list(DEFAULT_METHOD),
        help="LCIA method tuple",
    )
    p_e2e.set_defaults(func=_cmd_end_to_end)

    return parser


def _ask(msg: str, default: str) -> str:
    raw = input(
        f"{CLR_PROMPT}{msg}{CLR_RESET} "
        f"{CLR_DIM}[default: {default}]{CLR_RESET} "
        f"{CLR_INPUT}"
    ).strip()
    print(CLR_RESET, end="")
    return raw or default


def _ask_path(
    msg: str,
    default: Optional[Path] = None,
    must_exist: bool = False,
    must_be_file: bool = False,
) -> Path:
    while True:
        default_txt = str(default) if default is not None else "none"
        raw = input(
            f"{CLR_PROMPT}{msg}{CLR_RESET} "
            f"{CLR_DIM}[default: {default_txt}]{CLR_RESET} "
            f"{CLR_INPUT}"
        ).strip()
        print(CLR_RESET, end="")
        if not raw and default is None:
            print(f"{CLR_ERR}Path required.{CLR_RESET}")
            continue
        path = Path(raw).expanduser() if raw else default
        if must_exist and not path.exists():
            print(f"{CLR_ERR}Path not found:{CLR_RESET} {path}")
            continue
        if must_be_file and not path.is_file():
            print(f"{CLR_ERR}File not found:{CLR_RESET} {path}")
            continue
        return path


def _ask_number(msg: str, default: str, min_value: int, max_value: int) -> int:
    while True:
        raw = (
            input(
                f"{CLR_PROMPT}{msg}{CLR_RESET} "
                f"{CLR_DIM}[default: {default}]{CLR_RESET} "
                f"{CLR_INPUT}"
            ).strip()
            or default
        )
        print(CLR_RESET, end="")
        try:
            value = int(raw)
        except ValueError:
            print(f"{CLR_ERR}Please enter a valid number.{CLR_RESET}")
            continue
        if value < min_value or value > max_value:
            print(
                f"{CLR_ERR}Please enter a number between {min_value} and {max_value}.{CLR_RESET}"
            )
            continue
        return value


def _ask_yes_no(msg: str, default_yes: bool = True) -> bool:
    suffix = "[Y/n]" if default_yes else "[y/N]"
    raw = input(
        f"{CLR_PROMPT}{msg}{CLR_RESET} "
        f"{CLR_DIM}{suffix}:{CLR_RESET} "
        f"{CLR_INPUT}"
    ).strip().lower()
    print(CLR_RESET, end="")
    if not raw:
        return default_yes
    return raw in {"y", "yes"}


def _list_project_names(bw_dir: Path) -> List[str]:
    if bd is None:
        return []
    os.environ.setdefault("BRIGHTWAY2_DIR", str(bw_dir))
    os.environ.setdefault("BW2_DIR", str(bw_dir))
    out: List[str] = []
    for proj in bd.projects:
        name = getattr(proj, "name", None)
        if not name:
            s = str(proj)
            name = s.replace("Project: ", "", 1) if s.startswith("Project: ") else s
        out.append(name)
    return sorted(set(out), key=str.lower)


def _choose_project(bw_dir: Path, default_project: str) -> str:
    projects = _list_project_names(bw_dir)
    print(f"\n{CLR_TITLE}Select Brightway project{CLR_RESET}")
    if projects:
        for i, name in enumerate(projects, 1):
            marker = " (default)" if name == default_project else ""
            print(f"  {CLR_OPTION}{i}){CLR_RESET} {name}{marker}")
        print(
            f"  {CLR_OPTION}{len(projects) + 1}){CLR_RESET} Enter a new/custom project name"
        )
        default_idx = projects.index(default_project) + 1 if default_project in projects else 1
        sel = _ask_number("Project option", str(default_idx), 1, len(projects) + 1)
        if sel <= len(projects):
            return projects[sel - 1]
    else:
        print(f"{CLR_WARN}No projects listed; type one manually.{CLR_RESET}")
    return _ask("Project name", default_project)


def _print_step(title: str) -> None:
    print(f"\n{CLR_TITLE}{CLR_BOLD}=== {title} ==={CLR_RESET}")


def _interactive_main() -> None:
    while True:
        print(f"\n{CLR_TITLE}{CLR_BOLD}CE-RISE DPP terminal wizard{CLR_RESET}")
        print(
            f"  {CLR_OPTION}1){CLR_RESET} Excel -> DPP -> reference flow -> foreground -> LCIA "
            f"{CLR_DIM}(recommended){CLR_RESET}"
        )
        print(f"  {CLR_OPTION}2){CLR_RESET} Existing DPP + study -> foreground -> LCIA")
        print(f"  {CLR_OPTION}3){CLR_RESET} Build foreground only from existing DPP")
        print(f"  {CLR_OPTION}4){CLR_RESET} Exit")
        choice = _ask_number("Select option", "1", 1, 4)

        if choice == 4:
            return

        _print_step("Environment")
        bw_dir = _ask_path("Brightway data directory", DEFAULT_BW_DIR, must_exist=True)
        project = _choose_project(bw_dir, DEFAULT_BW_PROJECT)
        db_name = _ask("Foreground DB name", DEFAULT_DB_NAME)
        bonsai_db = _ask("BONSAI database name", DEFAULT_BONSAI_DB)
        biosphere_db = _ask("Biosphere database name", DEFAULT_BIOSPHERE_DB)

        try:
            dbs = project_databases(project=project, bw_dir=bw_dir)
            if bonsai_db in dbs:
                print(
                    f"{CLR_OK}Database '{bonsai_db}' already found in project '{project}'.{CLR_RESET}"
                )
            else:
                print(
                    f"{CLR_WARN}Database '{bonsai_db}' not found in project '{project}'.{CLR_RESET}"
                )
                if not _ask_yes_no(f"Import '{bonsai_db}' now?", default_yes=True):
                    print(
                        f"{CLR_WARN}Stopped: '{bonsai_db}' is required for this workflow.{CLR_RESET}"
                    )
                    return
                imported = import_bonsai_if_missing(
                    project=project,
                    bw_dir=bw_dir,
                    bonsai_db=bonsai_db,
                    biosphere_db=biosphere_db,
                    remote_project=DEFAULT_REMOTE_PROJECT,
                )
                if imported:
                    print(f"{CLR_OK}Imported '{bonsai_db}' into project '{project}'.{CLR_RESET}")
                else:
                    print(
                        f"{CLR_WARN}'{bonsai_db}' already present in project '{project}'.{CLR_RESET}"
                    )
            # Hard stop if still missing after check/import, before any run step.
            dbs_after = project_databases(project=project, bw_dir=bw_dir)
            if bonsai_db not in dbs_after:
                print(
                    f"{CLR_ERR}Stopped: database '{bonsai_db}' is still missing in project '{project}'.{CLR_RESET}"
                )
                return

            if choice == 1:
                _print_step("Excel -> DPP -> LCIA")
                workbook = _ask_path(
                    "Excel workbook path",
                    DEFAULT_WORKBOOK_INPUT,
                    must_exist=True,
                    must_be_file=True,
                )
                dpp_out = _ask_path("Output DPP JSON", DEFAULT_EXAMPLE_DPP, must_exist=False)
                study_out = _ask_path("Output study config JSON", DEFAULT_EXAMPLE_STUDY, must_exist=False)
                set_brightway_context(project=project, bw_dir=bw_dir)
                precheck_cancelled = False

                while True:
                    _print_step("Pre-check workbook before writing JSON")
                    dpp = excel_to_dpp(workbook)
                    schema = load_schema(SCHEMA_PATH)
                    validate_with_schema(dpp, schema)
                    validate_dpp_semantics(dpp)
                    try:
                        payload, unresolved = _build_foreground_strict_payload(
                            dpp=dpp,
                            db_name=db_name,
                            bonsai_db=bonsai_db,
                            biosphere_db=biosphere_db,
                        )
                    except Exception as pre_err:
                        print(f"{CLR_ERR}Pre-check failed:{CLR_RESET} {pre_err}")
                        print(
                            f"{CLR_WARN}Fix the Excel and re-run the check. "
                            f"No JSON was written.{CLR_RESET}"
                        )
                        print(f"  {CLR_OPTION}1){CLR_RESET} Re-check same workbook")
                        print(f"  {CLR_OPTION}2){CLR_RESET} Choose another workbook")
                        print(f"  {CLR_OPTION}3){CLR_RESET} Cancel to main menu")
                        pre_choice = _ask_number("Pre-check option", "1", 1, 3)
                        if pre_choice == 1:
                            input(
                                f"{CLR_PROMPT}Press Enter after updating the workbook...{CLR_RESET}"
                            )
                            continue
                        if pre_choice == 2:
                            workbook = _ask_path(
                                "Excel workbook path",
                                DEFAULT_WORKBOOK_INPUT,
                                must_exist=True,
                                must_be_file=True,
                            )
                            continue
                        precheck_cancelled = True
                        break
                    print(f"{CLR_OK}Pre-check passed. JSON files will now be written.{CLR_RESET}")
                    break

                if precheck_cancelled:
                    continue

                save_json(dpp_out, dpp)
                print(f"DPP JSON created: {dpp_out}")
                study = interactive_study_config_from_dpp(dpp)
                save_json(study_out, study)
                print(f"Study config saved: {study_out}")

                scaling, demand, info = compute_fu_scaling(dpp, study, db_name=db_name)
                write_foreground_database(payload, db_name=db_name, overwrite=True)
                lcia_result = run_lca(demand, method=DEFAULT_METHOD)
                result = {
                    "dppPath": str(dpp_out),
                    "studyPath": str(study_out),
                    "scaling": scaling,
                    "demand": {str(k): v for k, v in demand.items()},
                    "mappingInfo": info,
                    "unresolved": unresolved,
                    "foregroundDatasetCount": len(payload),
                    "lcia": lcia_result,
                }
                print(f"\n{CLR_TITLE}Run result{CLR_RESET}")
                print(json.dumps(result, indent=2))

            elif choice == 2:
                _print_step("DPP + Study -> LCIA")
                dpp_path = _ask_path("DPP JSON path", DEFAULT_EXAMPLE_DPP, must_exist=True)
                study_path = _ask_path("Study config JSON path", DEFAULT_EXAMPLE_STUDY, must_exist=True)
                args = argparse.Namespace(
                    dpp=dpp_path,
                    study=study_path,
                    schema=SCHEMA_PATH,
                    db_name=db_name,
                    project=project,
                    bw_dir=bw_dir,
                    bonsai_db=bonsai_db,
                    biosphere_db=biosphere_db,
                    cpa_mapping=DEFAULT_CPA_MAPPING,
                    method=list(DEFAULT_METHOD),
                )
                _cmd_run_lca(args)

            elif choice == 3:
                _print_step("Build foreground only")
                dpp_path = _ask_path("DPP JSON path", DEFAULT_EXAMPLE_DPP, must_exist=True)
                fg_out = _ask_path("Foreground output JSON", DEFAULT_EXAMPLE_FG, must_exist=False)
                args = argparse.Namespace(
                    dpp=dpp_path,
                    output=fg_out,
                    schema=SCHEMA_PATH,
                    db_name=db_name,
                    project=project,
                    bw_dir=bw_dir,
                    bonsai_db=bonsai_db,
                    biosphere_db=biosphere_db,
                    cpa_mapping=DEFAULT_CPA_MAPPING,
                    write_bw=True,
                    allow_unresolved=False,
                )
                _cmd_foreground(args)
        except Exception as exc:
            print(f"\n{CLR_ERR}{CLR_BOLD}ERROR:{CLR_RESET} {exc}")
            print(
                f"{CLR_WARN}No auto-fix was applied. Please correct input and run again.{CLR_RESET}"
            )

        if not _ask_yes_no("\nRun another operation?", default_yes=True):
            return


def main() -> None:
    parser = build_arg_parser()
    if len(sys.argv) == 1:
        _interactive_main()
        return
    args = parser.parse_args()
    if not getattr(args, "cmd", None):
        parser.print_help()
        return
    args.func(args)


if __name__ == "__main__":
    main()
